from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Ошибка: Модуль openpyxl не установлен. Установите его: pip install openpyxl")
    raise

from storage import build_export_path


logger = logging.getLogger(__name__)

# Общая тонкая рамка для всех листов
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_header(ws, headers: List[str], fill_color: str) -> None:
    """Пишет строку заголовков и оформляет её."""
    ws.append(headers)

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _finalize_sheet(ws, headers: List[str], column_widths: List[int], centered: bool = False) -> None:
    """Проставляет ширину колонок, стили данных и закрепляет шапку."""
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    data_font = Font(size=11)
    alignment = (
        Alignment(horizontal="center", vertical="center")
        if centered
        else Alignment(vertical="center", wrap_text=True)
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = data_font
            cell.border = THIN_BORDER
            cell.alignment = alignment

    ws.freeze_panes = "A2"


def create_excel_export(
    company_news: Dict[str, List[Dict]],
    all_companies: List[Dict],
    output_path: Optional[Path] = None,
    all_news: Optional[List[Dict]] = None,
) -> str:
    """
    Создает Excel файл с листами:
    1. Все новости (исходный список до фильтрации по компаниям)
    2. Компании (все компании из базы данных)
    3. Новости по компаниям (анализ новостей на предмет упоминаний компаний)

    Возвращает путь к созданному файлу.
    """
    if output_path is None:
        output_path = build_export_path("xlsx")

    # Создаем новую книгу Excel
    wb = Workbook()

    create_all_news_sheet(wb, all_news or [])
    create_companies_sheet(wb, all_companies)
    create_company_news_sheet(wb, company_news)

    # Удаляем дефолтный пустой лист, созданный вместе с книгой
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1:
        del wb["Sheet"]

    # Сохраняем файл
    wb.save(output_path)
    
    logger.info(f"Excel файл успешно создан: {output_path}")
    return str(output_path)


def create_all_news_sheet(wb: Workbook, all_news: List[Dict]) -> None:
    """
    Создает лист со всем собранным списком новостей — до фильтрации по компаниям.

    Колонка «Найденные компании» показывает результат распознавания: пустая
    ячейка означает, что новость не отнесена ни к одной отслеживаемой компании.
    """
    ws = wb.create_sheet(title="Все новости")

    headers = [
        "Заголовок",
        "Источник",
        "Дата публикации",
        "Тип материала",
        "Теги",
        "Влияние на кадры",
        "Почему",
        "Ссылка",
        "Найденные компании",
        "Кол-во компаний",
        "Описание",
    ]
    _write_header(ws, headers, "1F6F45")

    for news in all_news:
        ws.append([
            news.get('title', ''),
            news.get('source', ''),
            news.get('published_at', ''),
            news.get('material_type', ''),
            news.get('tags', ''),
            news.get('hr_signal', ''),
            news.get('hr_reason', ''),
            news.get('url', ''),
            news.get('companies', ''),
            news.get('companies_count', 0),
            # `or ''` обязателен: описание может быть None, а None[:500] падает
            (news.get('summary') or '')[:500],
        ])

    _finalize_sheet(ws, headers, [60, 15, 20, 16, 30, 18, 45, 50, 35, 15, 70])


def create_companies_sheet(wb: Workbook, companies: List[Dict]) -> None:
    """
    Создает лист с информацией о компаниях.
    """
    ws = wb.create_sheet(title="Компании")

    headers = ["Название компании", "Источник", "Рейтинг", "Год", "Ссылка", "Дата добавления"]
    _write_header(ws, headers, "366092")

    for company in companies:
        ws.append([
            company.get('name', ''),
            company.get('source', ''),
            company.get('rank', ''),
            company.get('year', ''),
            company.get('url', ''),
            company.get('created_at', ''),
        ])

    _finalize_sheet(ws, headers, [35, 15, 10, 10, 50, 20])


def create_company_news_sheet(wb: Workbook, company_news: Dict[str, List[Dict]]) -> None:
    """
    Создает лист с новостями, сгруппированными по компаниям.
    """
    ws = wb.create_sheet(title="Новости по компаниям")
    
    # Заголовки
    headers = [
        "Компания",
        "Количество упоминаний",
        "Заголовок новости",
        "Тип материала",
        "Теги",
        "Влияние на кадры",
        "Почему",
        "Источник новости",
        "Дата публикации",
        "Ссылка",
        "Сниппет текста"
    ]
    _write_header(ws, headers, "C65911")

    # Добавляем данные
    row_index = 2
    for company_name, news_list in sorted(company_news.items()):
        for news in news_list:
            ws.append([
                company_name,
                news.get('mention_count', 0),
                news.get('title', ''),
                news.get('material_type', ''),
                news.get('tags', ''),
                news.get('hr_signal', ''),
                news.get('hr_reason', ''),
                news.get('source', ''),
                news.get('published_at', ''),
                news.get('url', ''),
                # `or ''` обязателен: сниппет может быть None, а None[:200] падает
                (news.get('text_snippet') or '')[:200],
            ])

            # Чередуем заливку строк для читаемости
            fill_color = "E2EFDA" if row_index % 2 == 0 else "DDEBF7"
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_index, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            row_index += 1

    _finalize_sheet(ws, headers, [25, 15, 50, 16, 30, 18, 45, 15, 20, 50, 60])


def export_analysis_to_excel(
    limit_news: int = 1000,
    start_date=None,
    end_date=None,
) -> str:
    """
    Основная функция для экспорта анализа в Excel.

    Обе границы периода вместе ограничивают выгрузку датой публикации; без них
    берутся последние `limit_news` записей.
    """
    from company_analyzer import (
        get_all_companies_from_db,
        get_news_and_companies_from_db,
        get_reference_companies,
    )

    logger.info("Начинаем анализ новостей на предмет упоминаний компаний...")
    all_news, company_news = get_news_and_companies_from_db(
        limit=limit_news, start_date=start_date, end_date=end_date
    )

    logger.info("Получаем список компаний из базы данных...")
    all_companies = get_all_companies_from_db()
    if not all_companies:
        # Таблица companies заполняется отдельной командой (cli.py seed-companies).
        # Пока она пуста, показываем справочный список отслеживаемых компаний,
        # чтобы лист «Компании» не состоял из одной шапки.
        logger.info("Таблица companies пуста — используем справочный список компаний")
        all_companies = get_reference_companies()

    logger.info("Создаем Excel файл...")
    excel_path = create_excel_export(company_news, all_companies, all_news=all_news)

    # Выводим статистику
    total_companies = len(company_news)
    total_news = sum(len(news_list) for news_list in company_news.values())

    logger.info(f"Анализ завершен:")
    logger.info(f"- Всего собрано новостей: {len(all_news)}")
    logger.info(f"- Обнаружено компаний: {total_companies}")
    logger.info(f"- Найдено новостей с упоминаниями: {total_news}")
    logger.info(f"- Всего компаний в базе: {len(all_companies)}")
    logger.info(f"- Excel файл сохранен: {excel_path}")
    
    return excel_path


if __name__ == "__main__":
    import sys
    
    # Настраиваем логирование
    logging.basicConfig(
        level=logging.INFO,
        format='[%(asctime)s] [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Определяем лимит новостей из аргументов командной строки
    limit = 1000
    if len(sys.argv) > 1:
        try:
            limit = int(sys.argv[1])
        except ValueError:
            print(f"Использование: python {sys.argv[0]} [лимит_новостей]")
            print(f"Пример: python {sys.argv[0]} 500")
            sys.exit(1)
    
    try:
        export_analysis_to_excel(limit_news=limit)
        print("\nЭкспорт успешно завершен!")
    except Exception as e:
        logger.error(f"Ошибка при экспорте: {e}")
        sys.exit(1)